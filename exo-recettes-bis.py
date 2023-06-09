from openpyxl import load_workbook

wb = load_workbook("Recettes Iris.xlsx")

ingredients = {}

for name in wb.sheetnames:
    ws = wb[name]

    for row in ws.iter_rows(4):
        if row[0].value is not None:
            name = row[0].value
            quantity = row[1].value
            unit = row[2].value

            data = ingredients.get(name, {"quantity": 0, "unit": unit})
            data["quantity"] += quantity
            ingredients[name] = data

        else:
            break

print(ingredients)




