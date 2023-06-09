from openpyxl import load_workbook

wb = load_workbook("Recettes Iris.xlsx")

ingredients = {}

for name in wb.sheetnames:
    ws = wb[name]

    for row in ws.iter_rows(4):
        if row[0].value is not None:
            name = row[0].value
            quantity = row[1].value
            unity = row[2].value

            if name not in ingredients:
                data = {"quantity": quantity, "unity": unity}
                ingredients[name] = data
            else:
                ingredients[name]["quantity"] += quantity
        else:
            break

print(ingredients)




